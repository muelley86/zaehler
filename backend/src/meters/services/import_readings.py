"""Parsen und Anlegen historischer Zählerstände aus Excel/CSV.

Erwartetes Tabellen-Layout (mit Nutzer geklärt): je **Zeile** eine Messstelle,
erste Spalte = MP-Name; je **Spalte** ein Monat, Spaltenüberschrift = konkretes
Ablese-Datum; jede Zelle = Zählerstand.

``build_preview`` parst die Datei und matcht MP-Namen automatisch (der Admin
korrigiert/ergänzt im Frontend). ``commit_readings`` legt die Readings für das
aufgelöste Mapping an — idempotent (bestehende ``(register_id, reading_at)``
werden übersprungen), Plausibilität wird nicht hart geblockt.
"""

from __future__ import annotations

import csv
import io
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from meters.core.config import settings
from meters.models import AuditAction, AuditEntityType, MeasuringPoint, Reading, Register
from meters.schemas.import_readings import (
    ImportCell,
    ImportCommitResponse,
    ImportCommitRow,
    ImportFailure,
    ImportPreviewResponse,
    ImportRow,
)
from meters.services.audit import record

_DATE_FORMATS = ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y", "%d/%m/%Y", "%m/%Y", "%Y-%m")


def _parse_text_date(text: str) -> date | None:
    text = text.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_header_date(value: object) -> date | None:
    """Spaltenüberschrift -> Datum. Excel liefert echte ``datetime``-Zellen;
    CSV/Text wird über gängige Formate geparst."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _parse_text_date(value)
    return None


def _parse_decimal(raw: str) -> Decimal:
    """Zahl aus Zelltext. Akzeptiert deutsches (1.234,5) und internationales
    (1234.5) Format. Wirft ``ValueError`` bei nicht-parsebarem Text."""
    s = raw.strip().replace(" ", "")
    if not s:
        raise ValueError("leer")
    if "," in s and "." in s:
        # Punkt = Tausender, Komma = Dezimal (deutsches Format).
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ValueError(f"keine Zahl: {raw!r}") from exc


def _cell_raw(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Excel liefert Zahlen als float; ganze Zahlen ohne ".0" anzeigen.
        return str(int(value)) if value.is_integer() else repr(value)
    return str(value)


def _rows_from_xlsx(content: bytes) -> list[list[object]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_from_csv(content: bytes) -> list[list[object]]:
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(row) for row in reader]


def _read_grid(filename: str, content: bytes) -> list[list[object]]:
    if filename.lower().endswith(".csv"):
        return _rows_from_csv(content)
    return _rows_from_xlsx(content)


def _name_lookup(db: Session) -> dict[str, int]:
    """case-insensitiv getrimmtes Name->MP-ID-Mapping. Doppelte Namen werden
    NICHT auto-gematcht (Eintrag entfernt), damit der Admin manuell entscheidet."""
    lookup: dict[str, int] = {}
    duplicates: set[str] = set()
    for mp in db.scalars(select(MeasuringPoint)):
        key = mp.name.strip().casefold()
        if key in lookup:
            duplicates.add(key)
        else:
            lookup[key] = mp.id
    for key in duplicates:
        lookup.pop(key, None)
    return lookup


def build_preview(db: Session, *, filename: str, content: bytes) -> ImportPreviewResponse:
    grid = _read_grid(filename, content)
    if not grid:
        return ImportPreviewResponse(reading_dates=[], rows=[], ignored_columns=[])

    header = grid[0]
    # Spalte 0 = MP-Name; ab Spalte 1 Datums-Spalten.
    date_cols: list[tuple[int, date]] = []
    ignored: list[str] = []
    for col_idx in range(1, len(header)):
        parsed = _coerce_header_date(header[col_idx])
        if parsed is None:
            label = _cell_raw(header[col_idx])
            if label:
                ignored.append(label)
        else:
            date_cols.append((col_idx, parsed))

    lookup = _name_lookup(db)
    rows: list[ImportRow] = []
    for row_idx, raw_row in enumerate(grid[1:], start=1):
        raw_name = _cell_raw(raw_row[0] if raw_row else "").strip()
        if not raw_name:
            continue  # Leerzeile überspringen
        matched = lookup.get(raw_name.casefold())
        cells: list[ImportCell] = []
        for col_idx, col_date in date_cols:
            raw_value = raw_row[col_idx] if col_idx < len(raw_row) else None
            raw_text = _cell_raw(raw_value)
            if not raw_text.strip():
                continue  # leere Zelle -> kein Wert, kein Fehler
            try:
                value = _parse_decimal(raw_text)
                cells.append(ImportCell(reading_date=col_date, raw=raw_text, value=value))
            except ValueError as exc:
                cells.append(ImportCell(reading_date=col_date, raw=raw_text, error=str(exc)))
        rows.append(ImportRow(index=row_idx, raw_name=raw_name, matched_mp_id=matched, cells=cells))

    return ImportPreviewResponse(
        reading_dates=[d for _, d in date_cols], rows=rows, ignored_columns=ignored
    )


def _reading_at(reading_date: date) -> datetime:
    """Historischer Monatswert -> Zeitstempel am **Tagesende** (23:59:59 lokal),
    als naive UTC gespeichert (wie alle Readings). Einheitlich mit dem Erfassen-
    Toggle „Historischer Monatswert" und der App-„Periodenende"-Konvention. Nicht
    Mitternacht -> kein Mitternacht-Shift; das lokale Kalenderdatum bleibt über
    consumption._local_date stabil."""
    local_eod = datetime.combine(reading_date, time(23, 59, 59), tzinfo=ZoneInfo(settings.timezone))
    return local_eod.astimezone(UTC).replace(tzinfo=None)


def commit_readings(
    db: Session,
    *,
    rows: list[ImportCommitRow],
    user_id: int,
    ip_address: str | None,
    source_filename: str | None,
) -> ImportCommitResponse:
    registers = {
        r.id: r
        for r in db.scalars(
            select(Register).where(Register.id.in_({row.register_id for row in rows}))
        )
    }

    failed: list[ImportFailure] = []
    intended: list[tuple[int, datetime, Decimal]] = []
    for row in rows:
        reg = registers.get(row.register_id)
        for cell in row.cells:
            if reg is None or not reg.is_active:
                failed.append(
                    ImportFailure(
                        register_id=row.register_id,
                        reading_date=cell.reading_date,
                        reason="Register nicht gefunden oder inaktiv",
                    )
                )
                continue
            intended.append((row.register_id, _reading_at(cell.reading_date), cell.value))

    # Innerhalb der Datei doppelte (Register, Zeitpunkt) -> nur einmal anlegen.
    seen: set[tuple[int, datetime]] = set()
    deduped: list[tuple[int, datetime, Decimal]] = []
    skipped = 0
    for reg_id, reading_at, value in intended:
        key = (reg_id, reading_at)
        if key in seen:
            skipped += 1
            continue
        seen.add(key)
        deduped.append((reg_id, reading_at, value))

    # Bereits in der DB vorhandene (Register, Zeitpunkt) überspringen.
    existing: set[tuple[int, datetime]] = set()
    if deduped:
        reg_ids = {it[0] for it in deduped}
        ats = {it[1] for it in deduped}
        existing = {
            (r.register_id, r.reading_at)
            for r in db.execute(
                select(Reading.register_id, Reading.reading_at).where(
                    Reading.register_id.in_(reg_ids), Reading.reading_at.in_(ats)
                )
            )
        }

    to_add: list[Reading] = []
    for reg_id, reading_at, value in deduped:
        if (reg_id, reading_at) in existing:
            skipped += 1
            continue
        to_add.append(
            Reading(
                register_id=reg_id, value=value, reading_at=reading_at, created_by_user_id=user_id
            )
        )
    db.add_all(to_add)
    db.flush()

    record(
        db,
        user_id=user_id,
        action=AuditAction.CREATE,
        entity_type=AuditEntityType.READING,
        entity_id=None,
        diff={
            "import": True,
            "created": len(to_add),
            "skipped_existing": skipped,
            "failed": len(failed),
            "source_filename": source_filename,
        },
        ip_address=ip_address,
    )
    return ImportCommitResponse(created=len(to_add), skipped_existing=skipped, failed=failed)
